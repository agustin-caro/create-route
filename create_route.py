import openpyxl
import sys

packs = list(sys.argv)
packs.pop(0)
n_packs = len(packs)
print(str(n_packs) + " packs -> "+ str(packs))

wb = openpyxl.load_workbook('base_model.xlsx')
ws = wb.active

if n_packs > 1:
    ws.insert_rows(4, n_packs-1)

for col in ws.iter_cols(min_row=3,max_col=31,max_row=3):
    for i in range(n_packs):
        indexCol = col[0].column
        ws.cell(col[0].row+i, indexCol).value = col[0].internal_value
        if indexCol == 30:
            ws.cell(col[0].row+i, indexCol).value = packs[i]

wb.save("routes.xlsx")
wb.close()

